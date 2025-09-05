import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_inventory_template(products_df: pd.DataFrame) -> BytesIO:
    """
    Create an Excel template for inventory management
    
    Args:
        products_df: DataFrame containing product information
    
    Returns:
        BytesIO object containing the Excel file
    """
    # Create a copy of the dataframe with required columns
    template_df = products_df[[
        'master_sku', 'playauto_sku', 'product_name', 
        'category', 'is_set', 'current_stock'
    ]].copy()
    
    # Rename columns to Korean
    template_df.columns = [
        '마스터 SKU', '플레이오토 SKU', '상품명', 
        '카테고리', '세트 유무', '현재 재고'
    ]
    
    # Add empty columns for user input
    template_df['입고량'] = 0
    template_df['출고량'] = 0
    template_df['비고'] = ''
    
    # Convert is_set boolean to Korean
    template_df['세트 유무'] = template_df['세트 유무'].map({True: '세트', False: '단품'})
    
    # Create Excel writer
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        template_df.to_excel(writer, sheet_name='재고관리', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['재고관리']
        
        # Apply formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adjust column widths
        column_widths = {
            'A': 15,  # 마스터 SKU
            'B': 15,  # 플레이오토 SKU
            'C': 30,  # 상품명
            'D': 15,  # 카테고리
            'E': 10,  # 세트 유무
            'F': 10,  # 현재 재고
            'G': 10,  # 입고량
            'H': 10,  # 출고량
            'I': 30,  # 비고
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        
        # Add data validation for 입고량 and 출고량 (positive integers only)
        from openpyxl.worksheet.datavalidation import DataValidation
        
        dv = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=True,
            errorTitle="잘못된 입력",
            error="0 이상의 정수를 입력해주세요."
        )
        
        # Apply to 입고량 and 출고량 columns
        dv.add(f'G2:H{len(template_df) + 1}')
        worksheet.add_data_validation(dv)
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    output.seek(0)
    return output

def validate_inventory_file(df: pd.DataFrame) -> Tuple[bool, Optional[str]]:
    """
    Validate uploaded inventory file
    
    Args:
        df: DataFrame from uploaded file
    
    Returns:
        Tuple of (is_valid, error_message)
    """
    # Check required columns
    required_columns = ['마스터 SKU', '입고량', '출고량']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        return False, f"필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}"
    
    # Check data types
    try:
        df['입고량'] = pd.to_numeric(df['입고량'], errors='coerce')
        df['출고량'] = pd.to_numeric(df['출고량'], errors='coerce')
    except Exception as e:
        return False, f"숫자 변환 오류: {str(e)}"
    
    # Check for negative values
    if (df['입고량'] < 0).any() or (df['출고량'] < 0).any():
        return False, "입고량과 출고량은 0 이상이어야 합니다."
    
    # Check for NaN values in quantity columns
    if df['입고량'].isna().all() and df['출고량'].isna().all():
        return False, "입고량 또는 출고량을 입력해주세요."
    
    return True, None

def process_inventory_upload(df: pd.DataFrame, products_dict: Dict[str, int]) -> List[Dict]:
    """
    Process uploaded inventory file and prepare transactions
    
    Args:
        df: Validated DataFrame from upload
        products_dict: Dictionary mapping master_sku to product_id
    
    Returns:
        List of transaction dictionaries
    """
    transactions = []
    
    for _, row in df.iterrows():
        master_sku = row['마스터 SKU']
        
        # Skip if product not found
        if master_sku not in products_dict:
            st.warning(f"제품을 찾을 수 없습니다: {master_sku}")
            continue
        
        product_id = products_dict[master_sku]
        
        # Process IN transaction
        in_qty = row.get('입고량', 0)
        if pd.notna(in_qty) and in_qty > 0:
            transactions.append({
                'product_id': product_id,
                'transaction_type': 'IN',
                'quantity': int(in_qty),
                'notes': row.get('비고', '')
            })
        
        # Process OUT transaction
        out_qty = row.get('출고량', 0)
        if pd.notna(out_qty) and out_qty > 0:
            transactions.append({
                'product_id': product_id,
                'transaction_type': 'OUT',
                'quantity': -int(out_qty),  # Negative for OUT
                'notes': row.get('비고', '')
            })
    
    return transactions

def create_order_sheet(orders_df: pd.DataFrame) -> BytesIO:
    """
    Create an order sheet Excel file
    
    Args:
        orders_df: DataFrame containing order recommendations
    
    Returns:
        BytesIO object containing the Excel file
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        orders_df.to_excel(writer, sheet_name='발주서', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['발주서']
        
        # Apply formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary row
        last_row = len(orders_df) + 2
        worksheet[f'A{last_row}'] = '합계'
        worksheet[f'A{last_row}'].font = Font(bold=True)
        
        # Add date
        worksheet[f'A{last_row + 2}'] = f'발주일: {datetime.now().strftime("%Y-%m-%d")}'
    
    output.seek(0)
    return output

def parse_sales_history_file(file, file_type: str = 'auto') -> pd.DataFrame:
    """
    Parse sales history file with various formats
    
    Args:
        file: Uploaded file object
        file_type: File type ('auto', 'csv', 'excel')
    
    Returns:
        Parsed DataFrame with standardized columns
    """
    if file_type == 'auto':
        file_type = 'csv' if file.name.endswith('.csv') else 'excel'
    
    try:
        if file_type == 'csv':
            df = pd.read_csv(file, encoding='utf-8-sig')
        else:
            # Try to read all sheets
            excel_file = pd.ExcelFile(file)
            dfs = []
            
            for sheet_name in excel_file.sheet_names:
                sheet_df = pd.read_excel(file, sheet_name=sheet_name)
                sheet_df['sheet_name'] = sheet_name
                dfs.append(sheet_df)
            
            df = pd.concat(dfs, ignore_index=True)
    
    except Exception as e:
        st.error(f"파일 읽기 오류: {str(e)}")
        return pd.DataFrame()
    
    # Standardize column names (customize based on your actual data format)
    column_mapping = {
        '날짜': 'date',
        '일자': 'date',
        '제품명': 'product_name',
        '상품명': 'product_name',
        'SKU': 'sku',
        '수량': 'quantity',
        '출고량': 'quantity',
        '판매량': 'quantity'
    }
    
    df.rename(columns=column_mapping, inplace=True)
    
    # Convert date column
    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
    
    # Convert quantity to numeric
    if 'quantity' in df.columns:
        df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce')
    
    return df